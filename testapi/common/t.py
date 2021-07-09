# coding=utf-8
from openpyxl import load_workbook


wb = load_workbook('../fileData/api测试用例.xlsx')  # 打开指定文件
sheet = wb['get_case']  # 打开指定sheet页
row = sheet.max_row  # 获取最大行
col = sheet.max_column  # 获取最大列
# sheet.cell(row=1,column=1,value='111')
sheet.cell(1,1).value='sss'
wb.save('../fileData/api测试用例.xlsx')