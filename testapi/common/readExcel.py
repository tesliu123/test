# coding=utf-8
import xlrd
from openpyxl import load_workbook


class ReadExcel:
    def get_excel(self, sheet_name):
        work = xlrd.open_workbook('../fileData/api测试用例.xlsx')
        sheet = work.sheet_by_name(sheet_name)
        max_rows = sheet.nrows
        max_cols = sheet.ncols
        all_result = []
        for r in range(1, max_rows):
            result = {}
            for c in range(max_cols):
                title = sheet.cell_value(0, c)
                value = sheet.cell_value(r, c)
                result[title] = value
            all_result.append(result)
        return all_result

    def wr_excel(self, sheet_name, case_id, content, content1):
        wb = load_workbook('../fileData/api测试用例.xlsx')  # 打开指定文件
        sheet = wb[sheet_name]  # 打开指定sheet页
        row = sheet.max_row  # 获取最大行
        col = sheet.max_column  # 获取最大列
        max_c = 4
        sheet.cell(1, max_c + 1).value = '实际返回结果'
        sheet.cell(1, max_c + 2).value = '校验结果'
        print(row, col)
        sheet.cell(case_id + 1, 5, value=content)
        sheet.cell(case_id + 1, 6, value=content1)

        wb.save('../fileData/api测试用例.xlsx')


if __name__ == '__main__':
    r = ReadExcel()
    r.wr_excel('get_case', 1, '文本', 'Pass')
